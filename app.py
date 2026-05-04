"""
Kniha jízd - generátor v2
Streamlit app pro automatizaci tvorby měsíční knihy jízd
"""
import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# KOMPLETNÍ SEZNAM DESTINACÍ z reálného seznamu poboček
# Formát: nazev: (km, cas_min, typ)
# ============================================================
DESTINACE = {
    # === Domov / kancelář ===
    "Praha 7 Jankovcova": (0, 0, "domov"),

    # === Praha pobočky ===
    "Praha 4 - Háje": (26, 40, "pobocka"),
    "Praha 4, Budějovická": (10, 25, "pobocka"),
    "Praha 5 - Anděl": (8, 25, "pobocka"),
    "Praha 6 - Dejvice": (7, 15, "pobocka"),
    "Praha 7 - Holešovice": (2, 10, "pobocka"),
    "Praha 9 - Čakovice": (10, 15, "pobocka"),
    "Praha 9 - Horní Počernice": (18, 30,),

    # === Středočeský kraj ===
    "Benešov": (60, 50, "pobocka"),
    "Beroun": (40, 40, "pobocka"),
    "Chrášťany": (20, 30, "pobocka_sklad"),
    "Kladno": (31, 40, "pobocka"),
    "Kolín": (70, 60, "pobocka"),
    "Mělník": (35, 40, "pobocka"),
    "Mladá Boleslav": (60, 60, "pobocka"),
    "Příbram": (50, 55, "pobocka"),
    "Poděbrady": (75, 60,),
    "Úžice": (25, 25, "sklad"),
    "Zdiby": (13, 15, "sklad"),

    # === Severní/Západní Čechy ===
    "Děčín,": (110, 90, "pobocka"),
    "Chomutov": (98, 80, "pobocka"),
    "Karlovy Vary": (130, 120, "pobocka"),
    "Liberec": (110, 80, "pobocka"),
    "Most": (95, 75, "pobocka"),
    "Plzeň": (98, 75, "pobocka"),
    "Ústí nad Labem": (86, 60, "pobocka"),

    # === Jižní Čechy + Vysočina ===
    "České Budějovice": (162, 120, "pobocka"),
    "Jihlava": (140, 105, "pobocka"),
    "Tábor": (100, 105, "pobocka"),
    "Třebíč": (180, 120, "pobocka"),

    # === Východní Čechy ===
    "Hradec Králové": (111, 80, "pobocka"),
    "Pardubice": (122, 90, "pobocka"),

    # === Morava ===
    "Brno": (221, 150, "pobocka"),
    "Frýdek-Místek": (382, 255, "pobocka"),
    "Olomouc": (287, 210, "pobocka"),
    "Opava": (380, 260, "pobocka"),
    "Ostrava": (380, 250, "pobocka"),
    "Prostějov": (270, 195, "pobocka"),
    "Zlín": (303, 210, "pobocka"),

    # === Slovensko ===
    "Banská Bystrica": (550, 310, "pobocka"),
    "Bratislava": (350, 210, "pobocka"),
    "Košice": (681, 440, "pobocka"),
    "Poprad": (565, 380, "pobocka"),
    "Prešov": (651, 425, "pobocka"),
    "Senec": (350, 225, "pobocka_sklad"),
    "SKLC3": (368, 220, "sklad"),
    "Trenčín": (350, 245, "pobocka"),
    "Trnava": (402, 235, "pobocka"),
    "Žilina": (430, 285, "pobocka"),

    # === Maďarsko ===
    "Budapešť centrála (CBP), Róbert Károly": (548, 335, "pobocka_sklad"),
    "Budapest Blaha, József krt": (542, 335, "pobocka"),
    "Budapest Újbuda, Fehérvári út": (539, 325, "pobocka"),
    "HULC1 Budapest Sziget, Szigetszentmiklós": (548, 335, "pobocka_sklad"),

    # === Rakousko ===
    "Wien Karlsplatz, Getreidemarkt": (348, 225, "pobocka"),

    # === Pumpy a hotely (časté zastávky) ===
    "Praha 4 - Olbrachtova (pumpa)": (8, 20, "pumpa"),
    "Praha - Horní Počernice II (pumpa)": (15, 25, "pumpa"),
    "Chotýčany D3 (pumpa)": (130, 90, "pumpa"),
    "Zelenáč D2 (pumpa, směr BA)": (270, 165, "pumpa"),
    "Bratislava - hotel": (350, 210, "hotel"),
}

# Křížové vzdálenosti mezi destinacemi
KRIZOVE_VZDALENOSTI = {
    # SK
    ("SKLC3, Šákoňská cesta", "Bratislava, Mlynské nivy"): 18,
    ("Bratislava, Mlynské nivy", "SKLC3, Šákoňská cesta"): 18,
    ("SKLC3, Šákoňská cesta", "Bratislava - hotel"): 18,
    ("Bratislava - hotel", "SKLC3, Šákoňská cesta"): 18,
    ("SKLC3, Šákoňská cesta", "Bratislava Petržalka, Rusovská cesta"): 22,
    ("SKLC3, Šákoňská cesta", "Trnava, Starohájska"): 50,
    ("Trnava, Starohájska", "SKLC3, Šákoňská cesta"): 50,
    ("SKLC3, Šákoňská cesta", "Brno - střed, Skořepka"): 145,
    ("Brno - střed, Skořepka", "SKLC3, Šákoňská cesta"): 145,
    ("Bratislava, Mlynské nivy", "Trnava, Starohájska"): 55,
    ("Bratislava, Mlynské nivy", "Nitra, Akademická"): 95,
    ("Trnava, Starohájska", "Nitra, Akademická"): 50,

    # CZ
    ("Hradec Králové, Hořická", "Pardubice, Palackého"): 26,
    ("Pardubice, Palackého", "Hradec Králové, Hořická"): 26,
    ("Hradec Králové, Hořická", "Olomouc, Nedvědova"): 175,
    ("Olomouc, Nedvědova", "Ostrava, Novinářská"): 95,
    ("Olomouc, Nedvědova", "Prostějov, Poděbradovo nám."): 18,
    ("Brno - střed, Skořepka", "Olomouc, Nedvědova"): 80,
    ("Brno - střed, Skořepka", "Jihlava, Chlumova"): 90,
    ("Brno - střed, Skořepka", "Zlín, Jana Antonína Bati"): 100,
    ("Plzeň, Truhlářská", "Karlovy Vary, Sokolovská"): 82,
    ("České Budějovice, Průběžná", "Plzeň, Truhlářská"): 134,

    # Pumpy
    ("Chotýčany D3 (pumpa)", "České Budějovice, Průběžná"): 30,
    ("České Budějovice, Průběžná", "Chotýčany D3 (pumpa)"): 30,
    ("Zelenáč D2 (pumpa, směr BA)", "SKLC3, Šákoňská cesta"): 100,
    ("Praha 9 - Horní Počernice", "Horní Počernice (pumpa)"): 17,
    ("Zelenáč D2 (pumpa, směr BA)", "Bratislava, Mlynské nivy"): 80,
}

UCEL_OPTIONS = [
    "Pracovní jednání",
    "Pracovní schůzka",
    "Pracovní cesta",
    "Schůzka",
    "Schůzka + tankování",
    "Audit pobočky",
    "Návrat",
    "Návrat + tankování",
    "Cesta + tankování",
    "Ubytování",
    "Sklad - kontrola",
    "Návštěva skladu",
]

# ============================================================
# FUNKCE
# ============================================================
def init_state():
    if "destinace" not in st.session_state:
        st.session_state.destinace = DESTINACE.copy()
    if "krizove" not in st.session_state:
        st.session_state.krizove = KRIZOVE_VZDALENOSTI.copy()
    if "tankovani" not in st.session_state:
        st.session_state.tankovani = []
    if "jizdy" not in st.session_state:
        st.session_state.jizdy = []
    if "stav_pocatek" not in st.session_state:
        st.session_state.stav_pocatek = 136030


def vzdalenost(odkud, kam):
    if odkud == kam:
        return 0
    if (odkud, kam) in st.session_state.krizove:
        return st.session_state.krizove[(odkud, kam)]
    if odkud == "Praha 7 Jankovcova":
        return st.session_state.destinace.get(kam, (0, 0, ""))[0]
    if kam == "Praha 7 Jankovcova":
        return st.session_state.destinace.get(odkud, (0, 0, ""))[0]
    od = st.session_state.destinace.get(odkud, (0, 0, ""))[0]
    do = st.session_state.destinace.get(kam, (0, 0, ""))[0]
    return abs(od - do)


def cas_jizdy(odkud, kam):
    if odkud == "Praha 7 Jankovcova":
        return st.session_state.destinace.get(kam, (0, 0, ""))[1]
    if kam == "Praha 7 Jankovcova":
        return st.session_state.destinace.get(odkud, (0, 0, ""))[1]
    return int(vzdalenost(odkud, kam) * 1.0)


def parsuj_tankovani_excel(uploaded_file):
    df = pd.read_excel(uploaded_file)
    df.columns = [str(c).strip() for c in df.columns]
    sloupce = {col.lower(): col for col in df.columns}

    def najdi(klic_castecne):
        for low, orig in sloupce.items():
            if klic_castecne in low:
                return orig
        return None

    col_datum = najdi("datum")
    col_km = najdi("stav km") or najdi("stav")
    col_litry = najdi("množství") or najdi("mnozstvi") or najdi("litry")
    col_misto = najdi("čerpací") or najdi("cerpaci") or najdi("pumpa") or najdi("stanic")
    col_cena = najdi("částka") or najdi("castka") or najdi("cena")

    tankovani = []
    for _, row in df.iterrows():
        try:
            datum = row[col_datum] if col_datum else None
            km = int(row[col_km]) if col_km and pd.notna(row[col_km]) else None
            litry = float(row[col_litry]) if col_litry and pd.notna(row[col_litry]) else None
            misto = str(row[col_misto]) if col_misto else ""
            cena = float(row[col_cena]) if col_cena and pd.notna(row[col_cena]) else None
            if km and litry:
                tankovani.append({
                    "datum": datum, "km": km, "litry": litry,
                    "misto": misto.strip(), "cena": cena,
                })
        except Exception:
            continue
    tankovani.sort(key=lambda x: x["km"])
    return tankovani


def export_excel(jizdy, mesic_rok, stav_pocatek):
    wb = Workbook()
    ws = wb.active
    ws.title = mesic_rok

    ws['A1'] = "Typ vozidla:"; ws['B1'] = "doplnit"
    ws['A2'] = "RZ (SPZ) vozidla:"; ws['B2'] = "5SD7805"
    ws['A3'] = "Palivo:"; ws['B3'] = "Diesel"
    ws['A4'] = "Spotřeba paliva dle VTP"; ws['B4'] = 3.7
    ws['A5'] = "Počáteční stav tachometru:"; ws['B5'] = stav_pocatek
    final_km = jizdy[-1]["km_konec"] if jizdy else stav_pocatek
    ws['A6'] = "Konečný stav tachometru:"; ws['B6'] = final_km
    ws['A7'] = "Soukromé km:"; ws['B7'] = 0

    for i in range(1, 8):
        ws[f'A{i}'].font = Font(name='Arial', bold=True, size=10)

    headers = ['Datum', 'Výchozí místo', 'Cílové místo', 'Účel cesty',
               'Čas odjezdu', 'Čas příjezdu', 'Druh jízdy',
               'Stav km na začátku', 'Stav km na konci',
               'Počet kilometrů', 'Tankování litry', 'Cena phm', 'Řidič']
    yellow_fill = PatternFill('solid', start_color='FFF2CC')
    header_fill = PatternFill('solid', start_color='D9D9D9')
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    arial_bold = Font(name='Arial', bold=True, size=10)
    arial = Font(name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    HEADER_ROW = 9
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=col, value=h)
        c.font = arial_bold; c.fill = header_fill
        c.alignment = center; c.border = border

    row = HEADER_ROW + 1
    for j in jizdy:
        values = [
            j["datum"], j["odkud"], j["kam"], j["ucel"],
            j.get("cas_odj", ""), j.get("cas_prij", ""), j["druh"],
            j["km_zacatek"], j["km_konec"], j["km_pocet"],
            j.get("litry"), j.get("cena"), j.get("ridic", "MaWy"),
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = arial; c.border = border
            if col == 10: c.fill = yellow_fill
            if col >= 5:
                c.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

    widths = [11, 28, 32, 26, 11, 11, 11, 16, 16, 13, 13, 11, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HEADER_ROW].height = 32

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def format_cas(minuty):
    h = minuty // 60
    m = minuty % 60
    return f"{h}:{m:02d}"


# ============================================================
# UI
# ============================================================
st.set_page_config(page_title="Kniha jízd", page_icon="📕", layout="wide")
init_state()

st.title("📕 Kniha jízd – generátor")
st.caption(f"Automatizace měsíčního zápisu • {len(st.session_state.destinace)} destinací v databázi")

# --- KROK 1
with st.expander("⚙️ Krok 1 – nastavení měsíce", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        mesic = st.selectbox("Měsíc",
            ["leden", "únor", "březen", "duben", "květen", "červen",
             "červenec", "srpen", "září", "říjen", "listopad", "prosinec"],
            index=3)
    with col2:
        rok = st.number_input("Rok", min_value=2024, max_value=2030, value=2026)
    with col3:
        st.session_state.stav_pocatek = st.number_input(
            "Počáteční stav tachometru (km)",
            min_value=0, value=st.session_state.stav_pocatek, step=1,
        )

# --- KROK 2
with st.expander("⛽ Krok 2 – výpis tankování", expanded=True):
    st.markdown("Nahraj Excel z Teams, nebo přidávej tankování ručně.")
    uploaded = st.file_uploader("Nahrát Excel", type=["xlsx", "xls"])
    if uploaded:
        try:
            st.session_state.tankovani = parsuj_tankovani_excel(uploaded)
            st.success(f"✓ Načteno {len(st.session_state.tankovani)} tankování")
        except Exception as e:
            st.error(f"Chyba: {e}")

    if st.button("➕ Přidat tankování ručně"):
        st.session_state.tankovani.append({
            "datum": date.today(), "km": 0, "litry": 0.0, "misto": "", "cena": 0.0,
        })

    if st.session_state.tankovani:
        st.markdown("**Tankování (lze upravit):**")
        for i, t in enumerate(st.session_state.tankovani):
            cols = st.columns([2, 3, 2, 2, 2, 1])
            with cols[0]:
                t["datum"] = st.date_input("Datum",
                    value=t["datum"] if isinstance(t["datum"], (date, datetime)) else date.today(),
                    key=f"t_datum_{i}", label_visibility="collapsed")
            with cols[1]:
                t["misto"] = st.text_input("Místo", value=t["misto"], key=f"t_misto_{i}", label_visibility="collapsed")
            with cols[2]:
                t["km"] = st.number_input("KM", value=int(t["km"]), key=f"t_km_{i}", label_visibility="collapsed")
            with cols[3]:
                t["litry"] = st.number_input("L", value=float(t["litry"]), step=0.01, key=f"t_litry_{i}", label_visibility="collapsed")
            with cols[4]:
                t["cena"] = st.number_input("Kč", value=float(t["cena"] or 0), step=0.01, key=f"t_cena_{i}", label_visibility="collapsed")
            with cols[5]:
                if st.button("🗑", key=f"t_del_{i}"):
                    st.session_state.tankovani.pop(i)
                    st.rerun()

# --- KROK 3
with st.expander("🚗 Krok 3 – jízdy", expanded=True):
    if not st.session_state.tankovani:
        st.info("Nejdřív nahraj nebo přidej tankování.")
    else:
        prev_km = st.session_state.stav_pocatek
        st.markdown("**Úseky mezi tankováními:**")
        for i, t in enumerate(st.session_state.tankovani):
            cil_km = t["km"] - prev_km
            zapsano = sum(j["km_pocet"] for j in st.session_state.jizdy
                          if prev_km <= j["km_zacatek"] < t["km"])
            chybi = cil_km - zapsano
            datum_str = t["datum"].strftime("%-d.%-m.") if isinstance(t["datum"], (date, datetime)) else str(t["datum"])
            if chybi == 0:
                st.success(f"✓ Úsek do {datum_str} {t['misto']} (cíl {cil_km:,} km): zapsáno {zapsano:,} km")
            elif chybi > 0:
                st.error(f"✗ Úsek do {datum_str} {t['misto']} (cíl {cil_km:,} km): zapsáno {zapsano:,} km – **chybí {chybi:,} km**")
            else:
                st.warning(f"⚠️ Úsek do {datum_str} {t['misto']} (cíl {cil_km:,} km): přebývá {-chybi:,} km")
            prev_km = t["km"]

        st.markdown("---")
        st.markdown("**Přidat novou jízdu:**")

        col_filter1, col_filter2 = st.columns([2, 2])
        with col_filter1:
            filtr_typ = st.multiselect(
                "Filtr typu",
                ["pobocka", "sklad", "pobocka_sklad", "pumpa", "domov", "hotel"],
                default=["pobocka", "sklad", "pobocka_sklad", "domov", "hotel"]
            )
        with col_filter2:
            search = st.text_input("Hledat", placeholder="Brno, Praha, BA...")

        destinace_filtrovane = [
            name for name, (km, t, typ) in st.session_state.destinace.items()
            if typ in filtr_typ and (not search or search.lower() in name.lower())
        ]
        if not destinace_filtrovane:
            destinace_filtrovane = list(st.session_state.destinace.keys())

        c1, c2, c3, c4 = st.columns([2, 3, 3, 2])
        with c1:
            nove_datum = st.date_input("Datum", value=date(rok, 4, 1), key="nove_datum")
        with c2:
            odkud_idx = 0
            if "Praha 7 Jankovcova" in destinace_filtrovane:
                odkud_idx = destinace_filtrovane.index("Praha 7 Jankovcova")
            odkud = st.selectbox("Odkud", destinace_filtrovane, key="nove_odkud", index=odkud_idx)
        with c3:
            kam = st.selectbox("Kam", destinace_filtrovane, key="nove_kam",
                               index=min(1, len(destinace_filtrovane)-1))
        with c4:
            spocteno = vzdalenost(odkud, kam)
            km_pocet = st.number_input("KM", value=spocteno, min_value=0, key="nove_km")

        c5, c6, c7, c8 = st.columns([2, 2, 2, 2])
        with c5:
            cas_odj = st.text_input("Čas odjezdu", value="8:00", key="nove_odj")
        with c6:
            cas_min = cas_jizdy(odkud, kam)
            try:
                h, m = map(int, cas_odj.split(":"))
                cas_prij_default = format_cas(h * 60 + m + cas_min)
            except Exception:
                cas_prij_default = "9:30"
            cas_prij = st.text_input("Čas příjezdu", value=cas_prij_default, key="nove_prij")
        with c7:
            ucel = st.selectbox("Účel", UCEL_OPTIONS, key="nove_ucel")
        with c8:
            druh = st.selectbox("Druh", ["služební", ""], key="nove_druh")

        if st.button("➕ Přidat jízdu", type="primary"):
            if st.session_state.jizdy:
                last_km = st.session_state.jizdy[-1]["km_konec"]
            else:
                last_km = st.session_state.stav_pocatek

            nova = {
                "datum": nove_datum.strftime("%-d.%-m.%Y"),
                "odkud": odkud, "kam": kam, "ucel": ucel,
                "cas_odj": cas_odj, "cas_prij": cas_prij, "druh": druh,
                "km_zacatek": last_km, "km_konec": last_km + km_pocet,
                "km_pocet": km_pocet,
            }
            for t in st.session_state.tankovani:
                if t["km"] == nova["km_konec"]:
                    nova["litry"] = t["litry"]
                    nova["cena"] = t["cena"]
                    break
            st.session_state.jizdy.append(nova)
            st.rerun()

        if st.session_state.jizdy:
            st.markdown("**Zapsané jízdy:**")
            df_view = pd.DataFrame(st.session_state.jizdy)
            st.dataframe(df_view, use_container_width=True, hide_index=True)
            colb1, colb2 = st.columns(2)
            with colb1:
                if st.button("🗑️ Smazat poslední"):
                    st.session_state.jizdy.pop()
                    st.rerun()
            with colb2:
                if st.button("🗑️ Smazat všechny"):
                    st.session_state.jizdy = []
                    st.rerun()

# --- Šablony
with st.expander("📋 Šablony cest", expanded=False):
    st.caption("Rychlé přidání typických scénářů. Datum si pak doupravíš.")
    sablony_datum = st.date_input("Datum začátku", value=date(rok, 4, 1), key="sab_datum")

    def pridej_jizdy(jizdy_list, zacatek_datum):
        last_km = st.session_state.jizdy[-1]["km_konec"] if st.session_state.jizdy else st.session_state.stav_pocatek
        for j in jizdy_list:
            datum = zacatek_datum + timedelta(days=j.get("den", 0))
            km_pocet = vzdalenost(j["odkud"], j["kam"])
            nova = {
                "datum": datum.strftime("%-d.%-m.%Y"),
                "odkud": j["odkud"], "kam": j["kam"], "ucel": j["ucel"],
                "cas_odj": j["cas_odj"], "cas_prij": j["cas_prij"],
                "druh": j.get("druh", "služební"),
                "km_zacatek": last_km, "km_konec": last_km + km_pocet,
                "km_pocet": km_pocet,
            }
            for t in st.session_state.tankovani:
                if t["km"] == nova["km_konec"]:
                    nova["litry"] = t["litry"]
                    nova["cena"] = t["cena"]
                    break
            st.session_state.jizdy.append(nova)
            last_km = nova["km_konec"]

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**🇸🇰 SKLC3 + přespání BA (2 dny)**")
        if st.button("Přidat"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "SKLC3, Šákoňská cesta",
                 "ucel": "Pracovní cesta", "cas_odj": "7:00", "cas_prij": "11:00", "den": 0},
                {"odkud": "SKLC3, Šákoňská cesta", "kam": "Bratislava - hotel",
                 "ucel": "Ubytování", "cas_odj": "17:00", "cas_prij": "17:30", "druh": "", "den": 0},
                {"odkud": "Bratislava - hotel", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "13:00", "cas_prij": "17:00", "druh": "", "den": 1},
            ], sablony_datum)
            st.rerun()

        st.markdown("**🇨🇿 1denní Brno**")
        if st.button("Přidat ", key="brno"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "Brno - střed, Skořepka",
                 "ucel": "Pracovní jednání", "cas_odj": "7:00", "cas_prij": "9:30", "den": 0},
                {"odkud": "Brno - střed, Skořepka", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "16:00", "cas_prij": "18:30", "druh": "", "den": 0},
            ], sablony_datum)
            st.rerun()

    with col2:
        st.markdown("**🇨🇿 1denní ČB**")
        if st.button("Přidat", key="cb"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "České Budějovice, Průběžná",
                 "ucel": "Pracovní jednání", "cas_odj": "8:00", "cas_prij": "10:00", "den": 0},
                {"odkud": "České Budějovice, Průběžná", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "15:00", "cas_prij": "17:00", "druh": "", "den": 0},
            ], sablony_datum)
            st.rerun()

        st.markdown("**🇸🇰 SKLC3 + Trnava + BA (2 dny)**")
        if st.button("Přidat ", key="sk_okruh"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "SKLC3, Šákoňská cesta",
                 "ucel": "Pracovní cesta", "cas_odj": "7:00", "cas_prij": "11:00", "den": 0},
                {"odkud": "SKLC3, Šákoňská cesta", "kam": "Trnava, Starohájska",
                 "ucel": "Schůzka u klienta", "cas_odj": "13:00", "cas_prij": "13:50", "den": 0},
                {"odkud": "Trnava, Starohájska", "kam": "Bratislava - hotel",
                 "ucel": "Ubytování", "cas_odj": "17:00", "cas_prij": "18:00", "druh": "", "den": 0},
                {"odkud": "Bratislava - hotel", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "13:00", "cas_prij": "17:00", "druh": "", "den": 1},
            ], sablony_datum)
            st.rerun()

    with col3:
        st.markdown("**🇨🇿 Velký okruh Morava (2 dny)**")
        if st.button("Přidat ", key="morava"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "Hradec Králové, Hořická",
                 "ucel": "Pracovní jednání", "cas_odj": "7:00", "cas_prij": "8:30", "den": 0},
                {"odkud": "Hradec Králové, Hořická", "kam": "Olomouc, Nedvědova",
                 "ucel": "Schůzka", "cas_odj": "10:00", "cas_prij": "13:00", "den": 0},
                {"odkud": "Olomouc, Nedvědova", "kam": "Ostrava, Novinářská",
                 "ucel": "Ubytování", "cas_odj": "15:00", "cas_prij": "16:30", "druh": "", "den": 0},
                {"odkud": "Ostrava, Novinářská", "kam": "Brno - střed, Skořepka",
                 "ucel": "Schůzka cestou", "cas_odj": "9:00", "cas_prij": "12:00", "den": 1},
                {"odkud": "Brno - střed, Skořepka", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "15:00", "cas_prij": "17:30", "druh": "", "den": 1},
            ], sablony_datum)
            st.rerun()

        st.markdown("**🏙️ Pražské pobočky**")
        if st.button("Přidat ", key="praha"):
            pridej_jizdy([
                {"odkud": "Praha 7 Jankovcova", "kam": "Praha 4, Budějovická",
                 "ucel": "Pracovní schůzka", "cas_odj": "9:00", "cas_prij": "9:25", "den": 0},
                {"odkud": "Praha 4, Budějovická", "kam": "Praha 5 - Anděl, Štefánikova",
                 "ucel": "Schůzka", "cas_odj": "11:00", "cas_prij": "11:20", "den": 0},
                {"odkud": "Praha 5 - Anděl, Štefánikova", "kam": "Praha 7 Jankovcova",
                 "ucel": "Návrat", "cas_odj": "13:00", "cas_prij": "13:20", "druh": "", "den": 0},
            ], sablony_datum)
            st.rerun()

# --- Souhrn
if st.session_state.jizdy:
    st.markdown("---")
    st.markdown("### 📊 Souhrn")
    col1, col2, col3, col4 = st.columns(4)
    najeto = sum(j["km_pocet"] for j in st.session_state.jizdy)
    litry = sum(t["litry"] for t in st.session_state.tankovani)
    cena = sum((t.get("cena") or 0) for t in st.session_state.tankovani)
    col1.metric("Najeto", f"{najeto:,} km".replace(",", " "))
    col2.metric("Spotřeba", f"{litry:.2f} l")
    col3.metric("Cena PHM", f"{cena:,.0f} Kč".replace(",", " "))
    col4.metric("Konečný stav", f"{st.session_state.stav_pocatek + najeto:,} km".replace(",", " "))

# --- Export
st.markdown("---")
st.markdown("### ⬇️ Krok 4 – stáhnout knihu jízd")
if st.session_state.jizdy:
    excel_buffer = export_excel(
        st.session_state.jizdy, f"{mesic} {rok}", st.session_state.stav_pocatek,
    )
    st.download_button(
        label="⬇️ Stáhnout Excel",
        data=excel_buffer,
        file_name=f"kniha_jizd_{mesic}_{rok}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
else:
    st.info("Přidej alespoň jednu jízdu.")

# --- Sidebar
with st.sidebar:
    st.markdown("### 📍 Databáze destinací")
    typy = {}
    for _, (_, _, typ) in st.session_state.destinace.items():
        typy[typ] = typy.get(typ, 0) + 1
    typ_labels = {"pobocka": "🏢 Pobočky", "sklad": "📦 Sklady",
                  "pobocka_sklad": "🏢📦 Pobočky+sklady", "pumpa": "⛽ Pumpy",
                  "domov": "🏠 Domov", "hotel": "🏨 Hotely"}
    for typ, count in sorted(typy.items()):
        st.text(f"{typ_labels.get(typ, typ)}: {count}")

    st.markdown("---")
    st.markdown("**Přidat destinaci:**")
    nove_misto = st.text_input("Název")
    nove_km = st.number_input("Vzdálenost (km)", min_value=0, value=0)
    nove_cas = st.number_input("Čas (min)", min_value=0, value=0)
    nove_typ = st.selectbox("Typ", ["pobocka", "sklad", "pobocka_sklad", "pumpa", "hotel"])
    if st.button("Přidat"):
        if nove_misto:
            st.session_state.destinace[nove_misto] = (nove_km, nove_cas, nove_typ)
            st.success(f"Přidáno: {nove_misto}")
            st.rerun()

    st.markdown("---")
    if st.checkbox("Zobrazit všechny destinace"):
        for misto, (km, cas, typ) in sorted(st.session_state.destinace.items(), key=lambda x: x[1][0]):
            st.text(f"{km:>4} km | {misto[:30]}")
